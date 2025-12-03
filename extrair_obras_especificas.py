#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para extrair todas as obras de catálogos de leilões com monitoramento de valores em tempo real
Extrai TODAS as obras (quadros e esculturas) de cada catálogo fornecido
"""

import sys
import re
import time
import threading
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))

from database.database import SessionLocal, engine, init_db
from database.models import Base, Obra, ScrapingSession
from src.iarremate_scraper import IArremateScraper
from src.leiloes_br_scraper import LeiloesBRScraper


# URLs dos catálogos de leilões (extrai TODAS as obras de cada catálogo)
URLS_CATALOGOS = {
    'iarremate': [
        'https://www.iarremate.com/vitor_braga/022'
    ],
    'leiloes_br': [
        'https://www.miguelsalles.com.br/catalogo.asp?Num=50849',
        'https://www.robertohaddad.lel.br/catalogo.asp?Num=57470&Nav=lista&Sec=Catalogo&Pag=1&Srt=0&dia=4'
    ]
}


class HistoricoValor:
    """Classe para armazenar histórico de valores de uma obra"""
    def __init__(self, obra_id: int, url: str):
        self.obra_id = obra_id
        self.url = url
        self.valores = []  # Lista de tuplas (valor, timestamp, numero_lance)
        self.lock = threading.Lock()
        self.ultimo_numero_lance = 0  # Rastreia o último número de lance
    
    def adicionar_valor(self, valor: str, numero_lance: int = None, timestamp: datetime = None):
        """Adiciona um valor ao histórico com número de lance"""
        if timestamp is None:
            timestamp = datetime.utcnow()
        
        with self.lock:
            # Verificar se o valor mudou
            if not self.valores or self.valores[-1][0] != valor:
                # Se não foi fornecido número de lance, incrementar automaticamente
                if numero_lance is None:
                    self.ultimo_numero_lance += 1
                    numero_lance = self.ultimo_numero_lance
                else:
                    # Se foi fornecido, atualizar o contador se for maior
                    if numero_lance > self.ultimo_numero_lance:
                        self.ultimo_numero_lance = numero_lance
                
                self.valores.append((valor, timestamp, numero_lance))
                return True
        return False
    
    def obter_ultimo_valor(self) -> Optional[Tuple[str, datetime, int]]:
        """Retorna o último valor registrado com número de lance"""
        with self.lock:
            if self.valores:
                return self.valores[-1]
        return None
    
    def obter_todos_valores(self) -> List[Tuple[str, datetime, int]]:
        """Retorna todos os valores do histórico com números de lance"""
        with self.lock:
            return self.valores.copy()
    
    def obter_proximo_numero_lance(self) -> int:
        """Retorna o próximo número de lance"""
        with self.lock:
            return self.ultimo_numero_lance + 1


class ExtratorObrasEspecificas:
    """Classe para extrair obras específicas com monitoramento"""
    
    def __init__(self):
        self.scraper_iarremate = IArremateScraper()
        self.scraper_leiloes_br = LeiloesBRScraper()
        self.historicos = {}  # {url: HistoricoValor}
        self.monitores_ativos = {}  # {url: thread}
        self.lock = threading.Lock()
        self.intervalo_monitoramento = 30  # Verificar a cada 30 segundos
        self.db_session = SessionLocal()  # Sessão do banco de dados para verificar duplicatas
        self.obras_ja_verificadas = set()  # Cache em memória de URLs já verificadas
        self.obras_puladas = 0  # Contador de obras puladas por já existirem
        self.obras_ids_banco = {}  # {url: obra_id} - Mapear URLs para IDs no banco
        
    def _extrair_visitas(self, soup: BeautifulSoup) -> str:
        """Extrai o número de visitas da obra"""
        try:
            # Estratégia 1: Buscar ícone de pessoa caminhando seguido de número
            # (comum no Roberto Haddad)
            # Buscar por ícones SVG ou imagens de pessoa/visita
            visita_elements = soup.find_all(['span', 'div', 'p'], 
                                         class_=lambda x: x and any(
                                             keyword in str(x).lower() 
                                             for keyword in ['visita', 'visit', 'view', 'views', 'pessoa', 'person']
                                         ))
            
            for elem in visita_elements:
                # Buscar número próximo ao elemento
                texto = elem.get_text(strip=True)
                match = re.search(r'(\d+)', texto)
                if match:
                    return match.group(1)
                
                # Buscar no próximo irmão ou elemento próximo
                next_elem = elem.find_next_sibling()
                if next_elem:
                    texto_next = next_elem.get_text(strip=True)
                    match = re.search(r'(\d+)', texto_next)
                    if match:
                        return match.group(1)
            
            # Estratégia 2: Buscar padrão "Visitas: 312" ou "312 Visita(s)"
            texto_completo = soup.get_text()
            match = re.search(r'Visitas?[:\s]*(\d+)', texto_completo, re.IGNORECASE)
            if match:
                return match.group(1)
            
            # Estratégia 3: Buscar número próximo a ícones (Roberto Haddad específico)
            # Buscar elementos que contenham ícones de pessoa
            icons = soup.find_all(['i', 'svg', 'img'], 
                                class_=lambda x: x and any(
                                    keyword in str(x).lower() 
                                    for keyword in ['person', 'user', 'visita', 'walk', 'pessoa']
                                ))
            for icon in icons:
                # Buscar número no elemento pai ou próximo
                parent = icon.parent
                if parent:
                    texto = parent.get_text(strip=True)
                    match = re.search(r'(\d+)', texto)
                    if match:
                        return match.group(1)
                
                # Buscar próximo elemento
                next_elem = icon.find_next_sibling()
                if next_elem:
                    texto = next_elem.get_text(strip=True)
                    match = re.search(r'(\d+)', texto)
                    if match:
                        return match.group(1)
            
            # Estratégia 4: Buscar em elementos com classes relacionadas
            visita_elements = soup.find_all(['div', 'span'], 
                                         class_=lambda x: x and any(
                                             keyword in str(x).lower() 
                                             for keyword in ['visita', 'visit', 'view', 'views']
                                         ))
            for elem in visita_elements:
                texto = elem.get_text(strip=True)
                match = re.search(r'(\d+)', texto)
                if match:
                    return match.group(1)
        except Exception as e:
            print(f"  ⚠️ Erro ao extrair visitas: {e}")
        return 'N/A'
    
    def _extrair_status_lote(self, soup: BeautifulSoup) -> str:
        """Extrai o status do lote (vendido, disponível, arrematado, etc.)"""
        try:
            # Estratégia 1: Buscar botão "Lote vendido" (Roberto Haddad específico)
            buttons = soup.find_all(['button', 'a', 'div'], 
                                  class_=lambda x: x and any(
                                      keyword in str(x).lower() 
                                      for keyword in ['vendido', 'sold', 'lote', 'lot', 'status']
                                  ))
            for button in buttons:
                texto = button.get_text(strip=True).lower()
                if 'vendido' in texto or 'sold' in texto:
                    return 'Vendido'
                if 'arrematado' in texto:
                    return 'Arrematado'
                if 'disponível' in texto or 'disponivel' in texto:
                    return 'Disponível'
            
            # Estratégia 2: Buscar por texto "Lote vendido" na página
            texto_completo = soup.get_text().lower()
            
            # Buscar por status conhecidos
            status_keywords = {
                'vendido': ['lote vendido', 'vendido', 'sold', 'arrematado', 'adquirido'],
                'disponível': ['disponível', 'disponivel', 'available', 'em leilão', 'em leilao'],
                'finalizado': ['finalizado', 'encerrado', 'closed', 'terminado'],
                'reservado': ['reservado', 'reserved']
            }
            
            for status, keywords in status_keywords.items():
                for keyword in keywords:
                    if keyword in texto_completo:
                        # Verificar se está próximo a "lote" ou "obra"
                        pos = texto_completo.find(keyword)
                        contexto = texto_completo[max(0, pos-50):pos+50]
                        if any(palavra in contexto for palavra in ['lote', 'obra', 'peça', 'peca', 'item']):
                            return status.capitalize()
            
            # Estratégia 3: Buscar em elementos específicos com classes de status
            status_elements = soup.find_all(['div', 'span', 'strong', 'p', 'button'], 
                                         class_=lambda x: x and any(
                                             keyword in str(x).lower() 
                                             for keyword in ['status', 'estado', 'situacao', 'situação', 'vendido', 'sold']
                                         ))
            for elem in status_elements:
                texto = elem.get_text(strip=True).lower()
                if 'vendido' in texto or 'sold' in texto:
                    return 'Vendido'
                if 'arrematado' in texto:
                    return 'Arrematado'
                for status, keywords in status_keywords.items():
                    for keyword in keywords:
                        if keyword in texto:
                            return status.capitalize()
            
            # Estratégia 4: Buscar "Valor de venda: --" pode indicar vendido
            if 'valor de venda' in texto_completo and '--' in texto_completo:
                # Verificar se há botão de vendido próximo
                valor_elements = soup.find_all(text=re.compile(r'Valor de venda', re.IGNORECASE))
                for elem in valor_elements:
                    parent = elem.parent
                    if parent:
                        # Buscar botão vendido próximo
                        proximo = parent.find_next(['button', 'div', 'span'], 
                                                  string=re.compile(r'vendido|sold', re.IGNORECASE))
                        if proximo:
                            return 'Vendido'
        except Exception as e:
            print(f"  ⚠️ Erro ao extrair status: {e}")
        return 'Disponível'  # Default se não encontrar
    
    def _verificar_se_eh_quadro_ou_escultura(self, soup: BeautifulSoup, titulo: str = '', categoria_card: str = '') -> bool:
        """Verifica se a obra é um quadro ou escultura"""
        try:
            # Verificar no título
            texto_busca = (titulo + ' ' + categoria_card).lower()
            
            # Palavras-chave de quadros
            keywords_quadros = ['quadro', 'pintura', 'painting', 'tela', 'canvas', 'óleo', 'oleo', 'aquarela', 
                               'desenho', 'drawing', 'gravura', 'litografia', 'serigrafia']
            
            # Palavras-chave de esculturas
            keywords_esculturas = ['escultura', 'sculpture', 'estátua', 'estatua', 'bronze', 'mármore', 'marmore',
                                  'madeira', 'cerâmica', 'ceramica', 'porcelana']
            
            # Verificar se contém palavras-chave de quadros ou esculturas
            if any(keyword in texto_busca for keyword in keywords_quadros):
                return True
            if any(keyword in texto_busca for keyword in keywords_esculturas):
                return True
            
            # Verificar na página completa
            texto_completo = soup.get_text().lower()
            
            # Buscar por categorias na página
            categoria_elements = soup.find_all(['div', 'span', 'a'], 
                                              class_=lambda x: x and any(
                                                  keyword in str(x).lower() 
                                                  for keyword in ['categoria', 'category', 'tipo', 'type']
                                              ))
            for elem in categoria_elements:
                texto = elem.get_text().lower()
                if any(keyword in texto for keyword in keywords_quadros + keywords_esculturas):
                    return True
            
            # Verificar em breadcrumbs ou navegação
            breadcrumbs = soup.find_all(['nav', 'div'], 
                                       class_=lambda x: x and 'breadcrumb' in str(x).lower())
            for breadcrumb in breadcrumbs:
                texto = breadcrumb.get_text().lower()
                if any(keyword in texto for keyword in keywords_quadros + keywords_esculturas):
                    return True
            
        except:
            pass
        
        # Se não conseguir determinar, retornar True para não filtrar (pode ser ajustado)
        return True
    
    def obra_ja_existe(self, url: str, scraper_name: str) -> bool:
        """Verifica se a obra já existe no banco de dados"""
        try:
            # Verificar cache em memória primeiro
            if url in self.obras_ja_verificadas:
                return True
            
            # Verificar no banco de dados
            obra_existente = self.db_session.query(Obra).filter(
                Obra.url == url,
                Obra.scraper_name == scraper_name
            ).first()
            
            existe = obra_existente is not None
            
            # Adicionar ao cache se existir
            if existe:
                self.obras_ja_verificadas.add(url)
            
            return existe
        except Exception as e:
            # Em caso de erro, retornar False para não bloquear a extração
            print(f"  ⚠️ Erro ao verificar se obra existe: {e}")
            return False
    
    def extrair_obra_iarremate(self, url: str) -> Optional[Dict]:
        """Extrai dados de uma obra específica do iArremate"""
        # Verificar se a obra já existe no banco de dados
        if self.obra_ja_existe(url, 'iarremate'):
            self.obras_puladas += 1
            return None  # Retorna None para indicar que foi pulada
        
        response = self.scraper_iarremate.fazer_requisicao(url)
        if not response:
            print(f"  ❌ Erro ao acessar URL")
            return None
        
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Extrair dados usando métodos do scraper
        titulo = self.scraper_iarremate.extrair_titulo_iarremate(soup)
        descricao = self.scraper_iarremate.extrair_descricao_iarremate(soup, titulo)
        nome_artista = self.scraper_iarremate.extrair_nome_artista(titulo, descricao)
        valor = self.scraper_iarremate.extrair_valor_iarremate(soup)
        # Tentar extrair valor atual (com lances) se disponível
        valor_atual, numero_lances = self._extrair_valor_atual_com_lances(soup, valor)
        lote = self.scraper_iarremate.extrair_lote_iarremate(soup)
        data_inicio_leilao = self.scraper_iarremate.extrair_data_inicio_leilao_iarremate(soup)
        
        # Extrair data final do leilão (se disponível)
        data_final_leilao = self._extrair_data_final_iarremate(soup)
        
        # Verificar se é quadro ou escultura
        if not self._verificar_se_eh_quadro_ou_escultura(soup, titulo):
            return None  # Ignorar se não for quadro ou escultura
        
        # Extrair visitas e status
        visitas = self._extrair_visitas(soup)
        status_lote = self._extrair_status_lote(soup)
        
        obra_data = {
            'url': url,
            'scraper': 'iarremate',
            'titulo': titulo or 'N/A',
            'descricao': descricao or 'N/A',
            'nome_artista': nome_artista or 'N/A',
            'valor': valor or 'N/A',
            'valor_atual': valor_atual or valor or 'N/A',
            'numero_lances': numero_lances,
            'lote': lote or 'N/A',
            'visitas': visitas,
            'status_lote': status_lote,
            'data_inicio_leilao': data_inicio_leilao or 'N/A',
            'data_final_leilao': data_final_leilao or 'N/A',
            'leiloeiro': 'Vitor Braga',
            'local': 'N/A',
            'categoria': 'Quadros',
            'status_leilao': self._determinar_status_leilao(data_inicio_leilao, data_final_leilao)
        }
        
        return obra_data
    
    def extrair_obra_leiloes_br(self, url: str) -> Optional[Dict]:
        """Extrai dados de uma obra específica do LeilõesBR"""
        # Verificar se a obra já existe no banco de dados
        if self.obra_ja_existe(url, 'leiloes_br'):
            self.obras_puladas += 1
            return None  # Retorna None para indicar que foi pulada
        
        try:
            response = self.scraper_leiloes_br.fazer_requisicao(url)
            if not response:
                return None
        except Exception as e:
            # Silenciar erros de timeout - já são logados pelo scraper
            return None
        
        url_final = response.url
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Extrair dados usando métodos do scraper
        titulo = self.scraper_leiloes_br.extrair_titulo_leiloes_br(soup, 'N/A')
        descricao = self.scraper_leiloes_br.extrair_descricao_leiloes_br(soup, titulo)
        nome_artista = self.scraper_leiloes_br.extrair_nome_artista(titulo, descricao)
        valor = self.scraper_leiloes_br.extrair_valor_leiloes_br(soup, 'N/A')
        # Tentar extrair valor atual (com lances) se disponível
        valor_atual, numero_lances = self._extrair_valor_atual_com_lances(soup, valor)
        lote = self.scraper_leiloes_br.extrair_lote_leiloes_br(soup, url_final)
        data_inicio_leilao = self.scraper_leiloes_br.extrair_data_inicio_leilao_leiloes_br(soup)
        data_leilao = self.scraper_leiloes_br.extrair_data_leilao_leiloes_br(soup)
        leiloeiro = self.scraper_leiloes_br.extrair_leiloeiro_leiloes_br(soup)
        local = self.scraper_leiloes_br.extrair_local_leiloes_br(soup)
        
        # Extrair data final do leilão (se disponível)
        data_final_leilao = self._extrair_data_final_leiloes_br(soup)
        
        # Determinar leiloeiro do URL se não encontrado
        if not leiloeiro or leiloeiro == 'N/A':
            if 'miguelsalles' in url:
                leiloeiro = 'Miguel Salles'
            elif 'robertohaddad' in url:
                leiloeiro = 'Roberto Haddad'
        
        # Verificar se é quadro ou escultura
        if not self._verificar_se_eh_quadro_ou_escultura(soup, titulo):
            return None  # Ignorar se não for quadro ou escultura
        
        # Extrair visitas e status
        visitas = self._extrair_visitas(soup)
        status_lote = self._extrair_status_lote(soup)
        
        obra_data = {
            'url': url_final,
            'url_original': url,
            'scraper': 'leiloes_br',
            'titulo': titulo or 'N/A',
            'descricao': descricao or 'N/A',
            'nome_artista': nome_artista or 'N/A',
            'valor': valor or 'N/A',
            'valor_atual': valor_atual or valor or 'N/A',
            'numero_lances': numero_lances,
            'lote': lote or 'N/A',
            'visitas': visitas,
            'status_lote': status_lote,
            'data_inicio_leilao': data_inicio_leilao or data_leilao or 'N/A',
            'data_final_leilao': data_final_leilao or 'N/A',
            'leiloeiro': leiloeiro or 'N/A',
            'local': local or 'N/A',
            'categoria': 'Quadros',
            'status_leilao': self._determinar_status_leilao(data_inicio_leilao or data_leilao, data_final_leilao)
        }
        
        return obra_data
    
    def _extrair_data_final_iarremate(self, soup: BeautifulSoup) -> str:
        """Extrai a data final do leilão do iArremate"""
        try:
            # Buscar por "Fim", "Término", "Encerramento"
            keywords = ['fim', 'término', 'encerramento', 'final', 'data final']
            for keyword in keywords:
                textos = soup.find_all(text=re.compile(keyword, re.IGNORECASE))
                for texto in textos:
                    parent = texto.parent
                    if parent:
                        texto_completo = parent.get_text()
                        # Padrão: DD/MM/YYYY HH:MM
                        match = re.search(r'(\d{2}/\d{2}/\d{4})\s+(\d{2}:\d{2})', texto_completo)
                        if match:
                            return f"{match.group(1)} {match.group(2)}"
        except:
            pass
        return 'N/A'
    
    def _extrair_data_final_leiloes_br(self, soup: BeautifulSoup) -> str:
        """Extrai a data final do leilão do LeilõesBR"""
        try:
            # Buscar por "Fim", "Término", "Encerramento", "Último dia"
            keywords = ['fim', 'término', 'encerramento', 'final', 'último dia', 'ultimo dia']
            for keyword in keywords:
                textos = soup.find_all(text=re.compile(keyword, re.IGNORECASE))
                for texto in textos:
                    parent = texto.parent
                    if parent:
                        texto_completo = parent.get_text()
                        # Padrão: DD/MM/YYYY - HHh ou DD/MM/YYYY HH:MM
                        match = re.search(r'(\d{2}/\d{2}/\d{4})\s*[-–]\s*(\d{1,2})h', texto_completo)
                        if match:
                            return f"{match.group(1)} {match.group(2)}:00"
                        
                        match = re.search(r'(\d{2}/\d{2}/\d{4})\s+(\d{2}:\d{2})', texto_completo)
                        if match:
                            return f"{match.group(1)} {match.group(2)}"
            
            # Buscar em seções de "Dias do Leilão" que podem ter múltiplas datas
            # Exemplo: "3º DIA - 4/12/2025 - 20:00"
            dias_sections = soup.find_all(['div', 'section', 'ul', 'li'], 
                                        class_=lambda x: x and ('dia' in str(x).lower() or 'dias' in str(x).lower()))
            todas_datas = []
            for section in dias_sections:
                texto = section.get_text()
                # Buscar padrão: "Xº DIA - DD/MM/YYYY - HH:MM" ou "Xº DIA - DD/MM/YYYY HH:MM"
                matches = list(re.finditer(r'(\d{1,2})[º°]\s*DIA\s*[-–]\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*[-–]?\s*(\d{1,2}):(\d{2})', texto, re.IGNORECASE))
                for match in matches:
                    dia = match.group(2)
                    mes = match.group(3)
                    ano = match.group(4)
                    hora = match.group(5)
                    minuto = match.group(6)
                    data_str = f"{dia}/{mes}/{ano} {hora}:{minuto}"
                    todas_datas.append((int(match.group(1)), data_str))  # (número do dia, data)
            
            # Se encontrou múltiplas datas, pegar a última (maior número do dia)
            if todas_datas:
                todas_datas.sort(key=lambda x: x[0], reverse=True)
                return todas_datas[0][1]
            
            # Se não encontrou no formato específico, buscar última data mencionada
            dias_sections = soup.find_all(['div', 'section'], 
                                        class_=lambda x: x and 'dia' in str(x).lower())
            for section in dias_sections:
                texto = section.get_text()
                matches = list(re.finditer(r'(\d{2}/\d{2}/\d{4})', texto))
                if matches:
                    return matches[-1].group(1)
        except Exception as e:
            print(f"  ⚠️ Erro ao extrair data final: {e}")
        return 'N/A'
    
    def _extrair_valor_atual_com_lances(self, soup: BeautifulSoup, valor_fallback: str) -> Tuple[str, int]:
        """Extrai o valor atual e número de lances (para leilões em andamento)
        Retorna: (valor, numero_lances)
        """
        try:
            # Estratégia 1: Buscar número de lances pelo ícone de martelo (Roberto Haddad)
            # Buscar ícones de martelo/hammer
            hammer_icons = soup.find_all(['i', 'svg', 'img', 'span'], 
                                       class_=lambda x: x and any(
                                           keyword in str(x).lower() 
                                           for keyword in ['hammer', 'martelo', 'lance', 'bid']
                                       ))
            numero_lances = 0
            for icon in hammer_icons:
                # Buscar número no elemento pai ou próximo
                parent = icon.parent
                if parent:
                    texto = parent.get_text(strip=True)
                    match = re.search(r'(\d+)', texto)
                    if match:
                        numero_lances = int(match.group(1))
                        break
                
                # Buscar próximo elemento
                next_elem = icon.find_next_sibling()
                if next_elem:
                    texto = next_elem.get_text(strip=True)
                    match = re.search(r'(\d+)', texto)
                    if match:
                        numero_lances = int(match.group(1))
                        break
            
            # Estratégia 2: Buscar por "Valor atual" com número de lances
            # Exemplo: "Valor atual: (12 Lance(s)) R$ 8,000.00"
            texto_completo = soup.get_text()
            
            # Padrão: "Valor atual" seguido de número de lances e valor
            match = re.search(
                r'Valor\s+atual[:\s]*\(?\s*(\d+)\s*Lance\(s\)\s*\)?\s*R\$\s*([\d.,]+)',
                texto_completo,
                re.IGNORECASE
            )
            if match:
                numero_lances = int(match.group(1))
                valor = match.group(2)
                return (valor, numero_lances)
            
            # Padrão alternativo: "Valor atual:" seguido de valor (sem número de lances)
            match = re.search(
                r'Valor\s+atual[:\s]+R\$\s*([\d.,]+)',
                texto_completo,
                re.IGNORECASE
            )
            if match:
                valor = match.group(1)
                # Se já encontrou número de lances pelo ícone, usar ele
                if numero_lances == 0:
                    # Tentar encontrar número de lances em outro lugar
                    match_lances = re.search(r'(\d+)\s*Lance\(s\)', texto_completo, re.IGNORECASE)
                    numero_lances = int(match_lances.group(1)) if match_lances else 0
                return (valor, numero_lances)
            
            # Estratégia 3: Buscar "Valor de venda" (Roberto Haddad)
            match = re.search(
                r'Valor\s+de\s+venda[:\s]+R\$\s*([\d.,]+)',
                texto_completo,
                re.IGNORECASE
            )
            if match:
                valor = match.group(1)
                return (valor, numero_lances if numero_lances > 0 else 0)
            
            # Estratégia 4: Buscar em elementos com texto "Valor atual"
            valor_atual_elements = soup.find_all(
                text=re.compile(r'Valor\s+atual|Valor de venda', re.IGNORECASE)
            )
            for element in valor_atual_elements:
                parent = element.parent
                if parent:
                    texto = parent.get_text()
                    match = re.search(r'R\$\s*([\d.,]+)', texto)
                    if match:
                        valor = match.group(1)
                        # Tentar encontrar número de lances próximo
                        if numero_lances == 0:
                            match_lances = re.search(r'(\d+)\s*Lance\(s\)', texto, re.IGNORECASE)
                            numero_lances = int(match_lances.group(1)) if match_lances else 0
                        return (valor, numero_lances)
        except Exception as e:
            print(f"  ⚠️ Erro ao extrair valor atual: {e}")
        
        # Se não encontrou número de lances, retornar valor_fallback com 0 lances
        return (valor_fallback if isinstance(valor_fallback, str) else str(valor_fallback), 0)
    
    def _determinar_status_leilao(self, data_inicio: str, data_final: str) -> str:
        """Determina o status do leilão: 'agendado', 'em_andamento', 'finalizado'"""
        agora = datetime.utcnow()
        
        # Tentar parsear data de início
        data_inicio_dt = self._parsear_data(data_inicio)
        data_final_dt = self._parsear_data(data_final)
        
        if data_inicio_dt and data_final_dt:
            if agora < data_inicio_dt:
                return 'agendado'
            elif data_inicio_dt <= agora <= data_final_dt:
                return 'em_andamento'
            else:
                return 'finalizado'
        elif data_inicio_dt:
            if agora < data_inicio_dt:
                return 'agendado'
            elif (agora - data_inicio_dt).total_seconds() < 86400:  # Menos de 24h
                return 'em_andamento'
            else:
                return 'finalizado'
        
        return 'desconhecido'
    
    def _parsear_data(self, data_str: str) -> Optional[datetime]:
        """Parseia uma string de data para datetime"""
        if not data_str or data_str == 'N/A' or data_str == 'nao tem':
            return None
        
        try:
            # Formato: DD/MM/YYYY HH:MM ou DD/MM/YYYY
            match = re.search(r'(\d{2})/(\d{2})/(\d{4})(?:\s+(\d{2}):(\d{2}))?', data_str)
            if match:
                dia, mes, ano = int(match.group(1)), int(match.group(2)), int(match.group(3))
                hora = int(match.group(4)) if match.group(4) else 0
                minuto = int(match.group(5)) if match.group(5) else 0
                return datetime(ano, mes, dia, hora, minuto)
        except:
            pass
        
        return None
    
    def iniciar_monitoramento(self, url: str, obra_data: Dict):
        """Inicia monitoramento de valores em tempo real para uma obra"""
        status = obra_data.get('status_leilao', 'desconhecido')
        
        # Monitorar se o leilão estiver agendado, em andamento ou desconhecido
        # (desconhecido pode ser um leilão ativo que não conseguimos determinar a data)
        if status == 'finalizado':
            print(f"  ℹ️ Leilão finalizado (status: {status}). Monitoramento não iniciado.")
            return
        
        # Se for agendado, verificar se está próximo (até 2 horas antes)
        if status == 'agendado':
            data_inicio = obra_data.get('data_inicio_leilao', '')
            data_inicio_dt = self._parsear_data(data_inicio)
            if data_inicio_dt:
                agora = datetime.utcnow()
                tempo_restante = (data_inicio_dt - agora).total_seconds()
                if tempo_restante > 7200:  # Mais de 2 horas
                    print(f"  ℹ️ Leilão agendado para {data_inicio}. Monitoramento iniciará quando estiver próximo.")
                    # Mesmo assim, iniciar monitoramento para detectar quando começar
                else:
                    print(f"  ✅ Leilão agendado para {data_inicio}. Iniciando monitoramento.")
        
        # Criar histórico se não existir
        if url not in self.historicos:
            self.historicos[url] = HistoricoValor(obra_data.get('id', 0), url)
        
        # Adicionar valor inicial
        valor_inicial = obra_data.get('valor_atual', obra_data.get('valor', 'N/A'))
        numero_lances_inicial = obra_data.get('numero_lances', 0)
        if valor_inicial and valor_inicial != 'N/A':
            self.historicos[url].adicionar_valor(valor_inicial, numero_lances_inicial)
        
        # Verificar se já está sendo monitorado
        with self.lock:
            if url in self.monitores_ativos:
                print(f"  ℹ️ Monitoramento já está ativo para {url}")
                return
            
            # Criar thread de monitoramento
            thread = threading.Thread(
                target=self._monitorar_obra,
                args=(url, obra_data),
                daemon=True
            )
            thread.start()
            
            self.monitores_ativos[url] = {
                'thread': thread,
                'parar': False,
                'obra_data': obra_data
            }
            
            print(f"  🚀 Monitoramento iniciado para {url}")
    
    def _monitorar_obra(self, url: str, obra_data: Dict):
        """Monitora uma obra verificando mudanças de valor"""
        scraper = None
        if obra_data['scraper'] == 'iarremate':
            scraper = self.scraper_iarremate
        else:
            scraper = self.scraper_leiloes_br
        
        print(f"  📊 Iniciando monitoramento de valores para {url}")
        
        while True:
            with self.lock:
                if url not in self.monitores_ativos:
                    break
                if self.monitores_ativos[url].get('parar', False):
                    break
            
            try:
                # Fazer requisição para verificar valor atual
                response = scraper.fazer_requisicao(url)
                if response:
                    soup = BeautifulSoup(response.text, 'html.parser')
                    
                    if obra_data['scraper'] == 'iarremate':
                        valor_base = self.scraper_iarremate.extrair_valor_iarremate(soup)
                        novo_valor, novo_numero_lances = self._extrair_valor_atual_com_lances(soup, valor_base)
                    else:
                        valor_base = self.scraper_leiloes_br.extrair_valor_leiloes_br(soup, 'N/A')
                        novo_valor, novo_numero_lances = self._extrair_valor_atual_com_lances(soup, valor_base)
                    
                    if novo_valor and novo_valor != 'N/A':
                        # Adicionar ao histórico se mudou
                        if url in self.historicos:
                            if self.historicos[url].adicionar_valor(novo_valor, novo_numero_lances):
                                lance_info = f" (Lance {novo_numero_lances})" if novo_numero_lances > 0 else ""
                                print(f"  🔔 Mudança de valor detectada: R$ {novo_valor}{lance_info} ({datetime.now().strftime('%H:%M:%S')})")
                                # Atualizar obra_data
                                obra_data['valor_atual'] = novo_valor
                                obra_data['numero_lances'] = novo_numero_lances
                                
                                # ATUALIZAR NO BANCO DE DADOS
                                self._atualizar_obra_no_banco(url, novo_valor, novo_numero_lances, obra_data.get('scraper', ''))
            except Exception as e:
                print(f"  ⚠️ Erro ao verificar valor: {e}")
            
            # Aguardar antes da próxima verificação
            time.sleep(self.intervalo_monitoramento)
        
        print(f"  ✅ Monitoramento finalizado para {url}")
    
    def _atualizar_obra_no_banco(self, url: str, novo_valor: str, numero_lances: int, scraper_name: str):
        """Atualiza o valor e número de lances da obra no banco de dados"""
        try:
            # Usar uma nova sessão para evitar problemas de thread
            db = SessionLocal()
            try:
                # Buscar obra pelo URL e scraper
                obra = db.query(Obra).filter(
                    Obra.url == url,
                    Obra.scraper_name == scraper_name
                ).first()
                
                if obra:
                    # Atualizar valores
                    obra.valor_atualizado = novo_valor
                    obra.numero_lances = numero_lances
                    obra.ultima_atualizacao = datetime.utcnow()
                    db.commit()
                    print(f"  💾 Banco atualizado: Obra ID {obra.id} - Valor: R$ {novo_valor} (Lance {numero_lances})")
                else:
                    # Se não encontrou, pode ser que a URL seja diferente (url_original)
                    url_original = url
                    obra = db.query(Obra).filter(
                        Obra.url_original == url_original,
                        Obra.scraper_name == scraper_name
                    ).first()
                    
                    if obra:
                        obra.valor_atualizado = novo_valor
                        obra.numero_lances = numero_lances
                        obra.ultima_atualizacao = datetime.utcnow()
                        db.commit()
                        print(f"  💾 Banco atualizado (por URL original): Obra ID {obra.id} - Valor: R$ {novo_valor} (Lance {numero_lances})")
            finally:
                db.close()
        except Exception as e:
            print(f"  ⚠️ Erro ao atualizar obra no banco: {e}")
            # Não interromper o monitoramento por erro no banco
    
    def parar_monitoramento(self, url: str):
        """Para o monitoramento de uma obra"""
        with self.lock:
            if url in self.monitores_ativos:
                self.monitores_ativos[url]['parar'] = True
                print(f"  🛑 Parando monitoramento para {url}")
    
    def extrair_obras_do_catalogo_iarremate(self, url_catalogo: str) -> List[Dict]:
        """Extrai TODAS as obras de um catálogo do iArremate"""
        obras = []
        print(f"\n[IARREMATE] Extraindo catálogo: {url_catalogo}")
        
        # Processar TODAS as páginas até não encontrar mais obras
        pagina = 1
        max_paginas = 200  # Limite de segurança aumentado
        obras_ja_processadas = set()  # Para evitar duplicatas
        paginas_sem_obras = 0  # Contador de páginas consecutivas sem obras
        
        while pagina <= max_paginas:
            try:
                if pagina == 1:
                    url = url_catalogo
                else:
                    url = f"{url_catalogo}/pg{pagina}"
                
                print(f"  📖 Processando página {pagina}...")
                response = self.scraper_iarremate.fazer_requisicao(url)
                if not response:
                    print(f"    ⚠️ Erro ao acessar página {pagina}. Parando.")
                    break
                
                soup = BeautifulSoup(response.text, 'html.parser')
                
                # Encontrar todas as obras na página
                links_obras = soup.find_all('a', href=True)
                obras_pagina = 0
                obras_ignoradas = 0
                tem_obras = False
                
                for link in links_obras:
                    href = link.get('href', '')
                    if href and ('/belas-artes/' in href or '/quadro' in href or '/pintura' in href or '/escultura' in href or '/vitor_braga/' in href):
                        tem_obras = True
                        # Normalizar URL
                        if not href.startswith('http'):
                            if href.startswith('/'):
                                href = urljoin(self.scraper_iarremate.base_url, href)
                            else:
                                href = urljoin(url_catalogo, href)
                        
                        # Evitar duplicatas
                        if href in obras_ja_processadas:
                            continue
                        obras_ja_processadas.add(href)
                        
                        try:
                            obra = self.extrair_obra_iarremate(href)
                            if obra:  # Só adiciona se for quadro ou escultura
                                obras.append(obra)
                                obras_pagina += 1
                            else:
                                obras_ignoradas += 1
                        except Exception as e:
                            print(f"    ⚠️ Erro ao extrair obra {href}: {e}")
                            continue
                
                if not tem_obras:
                    paginas_sem_obras += 1
                    if paginas_sem_obras >= 2:  # Se 2 páginas consecutivas sem obras, parar
                        print(f"    ℹ️ {paginas_sem_obras} páginas consecutivas sem obras. Finalizando.")
                        break
                else:
                    paginas_sem_obras = 0  # Resetar contador
                
                if obras_pagina > 0:
                    print(f"    ✅ Página {pagina}: {obras_pagina} obras extraídas" + 
                          (f", {obras_ignoradas} ignoradas (não são quadros/esculturas)" if obras_ignoradas > 0 else ""))
                elif pagina == 1:
                    print(f"    ⚠️ Nenhuma obra encontrada na primeira página. Verifique a URL.")
                    break
                
                pagina += 1
                time.sleep(0.5)  # Delay entre páginas
                
            except Exception as e:
                print(f"  ❌ Erro ao processar página {pagina}: {e}")
                # Tentar continuar para próxima página
                pagina += 1
                if pagina > max_paginas:
                    break
                continue
        
        total_paginas_processadas = pagina - 1
        print(f"\n  ✅ Total de páginas processadas: {total_paginas_processadas}")
        print(f"  ✅ Total de obras extraídas do catálogo: {len(obras)}")
        return obras
    
    def _encontrar_obras_catalogo_especifico(self, soup: BeautifulSoup, url_base: str) -> List[Dict]:
        """Encontra obras em catálogos específicos (Miguel Salles, Roberto Haddad)"""
        obras = []
        urls_encontradas = set()  # Para evitar duplicatas
        
        try:
            # ESTRATÉGIA PRINCIPAL: Buscar links diretos para peca.asp?ID=... (mais confiável)
            # Miguel Salles e Roberto Haddad usam este formato: peca.asp?ID=26942673&...
            links_pecas = soup.find_all('a', href=re.compile(r'peca\.asp\?ID=\d+', re.IGNORECASE))
            
            print(f"    🔍 Encontrados {len(links_pecas)} links do tipo peca.asp?ID=...")
            
            for link in links_pecas:
                href = link.get('href', '')
                if not href:
                    continue
                
                # Normalizar URL
                if href.startswith('/'):
                    href = urljoin(url_base, href)
                elif not href.startswith('http'):
                    # Se não começa com http, construir URL completa
                    if url_base.endswith('/'):
                        href = url_base + href
                    else:
                        # Verificar se href começa com ? ou &
                        if href.startswith('?'):
                            href = url_base + href
                        else:
                            href = url_base + '/' + href
                
                # Evitar duplicatas
                if href in urls_encontradas:
                    continue
                urls_encontradas.add(href)
                
                # NÃO extrair título aqui - será extraído na página individual da obra
                # Isso evita pegar títulos genéricos como "Lotes relacionados"
                obras.append({
                    'url': href,
                    'titulo': 'N/A'  # Será extraído depois na página individual
                })
            
            # Estratégia 2: Buscar por números de lote na página (Miguel Salles específico)
            # Padrão: "Lote:1", "Lote 1", "LOTE 448", etc.
            if not obras or len(obras) < 5:
                # Buscar por texto "Lote" seguido de número
                lote_elements = soup.find_all(text=re.compile(r'Lote\s*:?\s*\d+', re.IGNORECASE))
                for lote_text in lote_elements:
                    parent = lote_text.parent
                    if parent:
                        # Buscar link próximo ao lote (pode estar no mesmo elemento ou próximo)
                        link = parent.find('a', href=True)
                        if not link:
                            # Buscar no próximo elemento irmão
                            link = parent.find_next_sibling('a', href=True)
                        if not link:
                            # Buscar no elemento pai
                            parent_parent = parent.find_parent(['div', 'li', 'article', 'section', 'td', 'tr'])
                            if parent_parent:
                                link = parent_parent.find('a', href=True)
                        if not link:
                            # Buscar próximo elemento com link
                            next_elem = parent.find_next(['a', 'div', 'li'])
                            if next_elem:
                                link = next_elem.find('a', href=True) if next_elem.name != 'a' else next_elem
                        
                        if link and link.get('href'):
                            href = link.get('href', '')
                            if href and ('peca.asp' in href.lower() or 'catalogo.asp' in href.lower() or 'id=' in href.lower()):
                                # Normalizar URL
                                if href.startswith('/'):
                                    href = urljoin(url_base, href)
                                elif not href.startswith('http'):
                                    href = urljoin(url_base, '/' + href)
                                
                                if href not in [o.get('url') for o in obras]:
                                    titulo = link.get_text(strip=True)
                                    if not titulo or len(titulo) < 5:
                                        # Buscar título no contexto do lote
                                        contexto = parent.get_text(strip=True)
                                        linhas = contexto.split('\n')
                                        for linha in linhas:
                                            linha = linha.strip()
                                            if len(linha) > 10 and not re.match(r'^[\d\sR$.,Lote:]+$', linha, re.IGNORECASE):
                                                titulo = linha
                                                break
                                    
                                obras.append({
                                    'url': href,
                                    'titulo': titulo
                                })
        
        except Exception as e:
            print(f"    ⚠️ Erro ao encontrar obras na página: {e}")
        
        return obras
    
    def extrair_obras_do_catalogo_leiloes_br(self, url_catalogo: str) -> List[Dict]:
        """Extrai TODAS as obras de um catálogo do LeilõesBR"""
        obras = []
        print(f"\n[LEILÕES BR] Extraindo catálogo: {url_catalogo}")
        
        # Extrair URL base
        from urllib.parse import urlparse
        parsed = urlparse(url_catalogo)
        url_base = f"{parsed.scheme}://{parsed.netloc}"
        
        # Tentar descobrir total de páginas primeiro
        total_paginas_estimado = self._descobrir_total_paginas_leiloes_br(url_catalogo)
        if total_paginas_estimado > 1:
            print(f"  📄 Total de páginas estimado: {total_paginas_estimado}")
        
        pagina = 1
        max_paginas = 200  # Limite de segurança aumentado
        obras_ja_processadas = set()  # Para evitar duplicatas
        paginas_sem_obras = 0  # Contador de páginas consecutivas sem obras
        
        # Processar TODAS as páginas até não encontrar mais obras
        while pagina <= max_paginas:
            try:
                # Construir URL da página (diferentes formatos para diferentes sites)
                url = self._construir_url_pagina_leiloes_br(url_catalogo, pagina)
                
                print(f"  📖 Processando página {pagina}...")
                
                # Aumentar delay para evitar timeouts
                time.sleep(1.0)  # Delay antes de fazer requisição
                
                response = self.scraper_leiloes_br.fazer_requisicao(url)
                if not response:
                    print(f"    ⚠️ Erro ao acessar página {pagina}. Parando.")
                    paginas_sem_obras += 1
                    if paginas_sem_obras >= 2:
                        break
                    pagina += 1
                    continue
                
                soup = BeautifulSoup(response.text, 'html.parser')
                
                # Usar método específico para catálogos (Miguel Salles, Roberto Haddad)
                obras_pagina = self._encontrar_obras_catalogo_especifico(soup, url_base)
                
                # Se não encontrou, tentar método genérico
                if not obras_pagina or len(obras_pagina) == 0:
                    obras_pagina = self.scraper_leiloes_br._encontrar_obras_na_pagina(soup)
                
                if not obras_pagina or len(obras_pagina) == 0:
                    paginas_sem_obras += 1
                    if paginas_sem_obras >= 2:
                        print(f"    ℹ️ {paginas_sem_obras} páginas consecutivas sem obras. Finalizando.")
                        break
                    print(f"    ℹ️ Nenhuma obra encontrada na página {pagina}. Continuando...")
                    pagina += 1
                    continue
                else:
                    paginas_sem_obras = 0  # Resetar contador
                
                print(f"    📋 Encontradas {len(obras_pagina)} obras na página {pagina}")
                
                # Processar cada obra encontrada (com delay para evitar timeouts)
                obras_processadas = 0
                obras_ignoradas = 0
                obras_erro = 0
                
                for idx, obra_data in enumerate(obras_pagina):
                    try:
                        url_obra = obra_data.get('url', '')
                        if not url_obra:
                            continue
                        
                        # Evitar processar a mesma obra duas vezes
                        if url_obra in obras_ja_processadas:
                            continue
                        obras_ja_processadas.add(url_obra)
                        
                        # Delay entre requisições para evitar timeout
                        if idx > 0 and idx % 5 == 0:
                            time.sleep(2)  # Delay maior a cada 5 obras
                        else:
                            time.sleep(0.8)  # Delay entre cada obra
                        
                        # Extrair dados completos da obra
                        obra = self.extrair_obra_leiloes_br(url_obra)
                        if obra:  # Só adiciona se for quadro ou escultura
                            obras.append(obra)
                            obras_processadas += 1
                        else:
                            obras_ignoradas += 1
                    except Exception as e:
                        obras_erro += 1
                        if obras_erro <= 3:  # Só mostrar primeiros 3 erros
                            print(f"    ⚠️ Erro ao extrair obra: {e}")
                        continue
                
                print(f"    ✅ Página {pagina}: {obras_processadas} obras extraídas" + 
                      (f", {obras_ignoradas} ignoradas (não são quadros/esculturas)" if obras_ignoradas > 0 else "") +
                      (f", {obras_erro} erros" if obras_erro > 0 else ""))
                
                # Se não processou nenhuma obra nova, pode ter chegado ao fim
                if obras_processadas == 0 and pagina > 1 and obras_erro == 0:
                    print(f"    ℹ️ Nenhuma obra nova na página {pagina}. Finalizando.")
                    break
                
                pagina += 1
                time.sleep(1.0)  # Delay entre páginas
                
            except Exception as e:
                print(f"  ❌ Erro ao processar página {pagina}: {e}")
                paginas_sem_obras += 1
                if paginas_sem_obras >= 2:
                    break
                # Tentar continuar para próxima página
                pagina += 1
                if pagina > max_paginas:
                    break
                time.sleep(2)  # Delay maior em caso de erro
                continue
        
        total_paginas_processadas = pagina - 1
        print(f"\n  ✅ Total de páginas processadas: {total_paginas_processadas}")
        print(f"  ✅ Total de obras extraídas do catálogo: {len(obras)}")
        return obras
    
    def _construir_url_pagina_leiloes_br(self, url_catalogo: str, pagina: int) -> str:
        """Constrói a URL da página específica do catálogo"""
        # Diferentes formatos de URL para diferentes sites
        if 'robertohaddad' in url_catalogo:
            # Roberto Haddad: ?Pag=1, ?Pag=2, etc. (já tem Pag no URL original)
            if pagina == 1:
                # Manter URL original na primeira página
                return url_catalogo
            else:
                # Remover parâmetro Pag existente se houver e adicionar novo
                url = re.sub(r'[&?]Pag=\d+', '', url_catalogo)
                if '?' in url:
                    return f"{url}&Pag={pagina}"
                else:
                    return f"{url}?Pag={pagina}"
        elif 'miguelsalles' in url_catalogo:
            # Miguel Salles: Página 1, 2, 3... (pode usar Pag ou outro parâmetro)
            if pagina == 1:
                # Remover qualquer parâmetro de página da URL original
                url = re.sub(r'[&?]Pag=\d+', '', url_catalogo)
                return url
            else:
                # Adicionar parâmetro Pag
                url = re.sub(r'[&?]Pag=\d+', '', url_catalogo)
                if '?' in url:
                    return f"{url}&Pag={pagina}"
                else:
                    return f"{url}?Pag={pagina}"
        else:
            # Formato genérico
            if pagina == 1:
                url = re.sub(r'[&?]Pag=\d+', '', url_catalogo)
                return url
            else:
                url = re.sub(r'[&?]Pag=\d+', '', url_catalogo)
                if '?' in url:
                    return f"{url}&Pag={pagina}"
                else:
                    return f"{url}?Pag={pagina}"
    
    def _descobrir_total_paginas_leiloes_br(self, url_catalogo: str) -> int:
        """Tenta descobrir o total de páginas do catálogo"""
        try:
            # Acessar primeira página
            url_primeira = self._construir_url_pagina_leiloes_br(url_catalogo, 1)
            response = self.scraper_leiloes_br.fazer_requisicao(url_primeira)
            if not response:
                return 1
            
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Buscar paginação na página
            # Estratégia 1: Buscar links de paginação
            links_pagina = soup.find_all('a', href=re.compile(r'[?&]Pag=\d+', re.IGNORECASE))
            max_pagina = 1
            for link in links_pagina:
                href = link.get('href', '')
                match = re.search(r'[?&]Pag=(\d+)', href, re.IGNORECASE)
                if match:
                    num = int(match.group(1))
                    if num > max_pagina:
                        max_pagina = num
            
            # Estratégia 2: Buscar texto "Página X de Y" ou similar
            texto = soup.get_text()
            match = re.search(r'página\s+\d+\s+de\s+(\d+)', texto, re.IGNORECASE)
            if match:
                num = int(match.group(1))
                if num > max_pagina:
                    max_pagina = num
            
            # Estratégia 3: Buscar números na paginação
            paginacao = soup.find_all(['div', 'nav', 'ul'], 
                                     class_=lambda x: x and any(
                                         keyword in str(x).lower() 
                                         for keyword in ['pagin', 'page', 'nav']
                                     ))
            for elem in paginacao:
                texto = elem.get_text()
                numeros = re.findall(r'\b(\d+)\b', texto)
                for num_str in numeros:
                    num = int(num_str)
                    if 1 < num < 1000 and num > max_pagina:
                        max_pagina = num
            
            return max_pagina if max_pagina > 1 else 1
        except:
            return 1
    
    def extrair_todas_obras(self) -> List[Dict]:
        """Extrai TODAS as obras de todos os catálogos"""
        obras = []
        
        print("=" * 80)
        print("EXTRAÇÃO DE TODAS AS OBRAS DOS CATÁLOGOS")
        print("Filtrando apenas: Quadros e Esculturas")
        print("=" * 80)
        
        # Extrair obras do iArremate
        for url_catalogo in URLS_CATALOGOS['iarremate']:
            try:
                obras_catalogo = self.extrair_obras_do_catalogo_iarremate(url_catalogo)
                obras.extend(obras_catalogo)
            except Exception as e:
                print(f"  ❌ Erro ao extrair catálogo iArremate {url_catalogo}: {e}")
        
        # Extrair obras do LeilõesBR
        for url_catalogo in URLS_CATALOGOS['leiloes_br']:
            try:
                obras_catalogo = self.extrair_obras_do_catalogo_leiloes_br(url_catalogo)
                obras.extend(obras_catalogo)
            except Exception as e:
                print(f"  ❌ Erro ao extrair catálogo LeilõesBR {url_catalogo}: {e}")
        
        print("\n" + "=" * 80)
        print(f"✅ Total de obras extraídas (Quadros e Esculturas): {len(obras)}")
        if self.obras_puladas > 0:
            print(f"⏭️  Total de obras puladas (já existentes no banco): {self.obras_puladas}")
        print("=" * 80)
        
        return obras
    
    def salvar_obras_no_banco(self, obras: List[Dict]):
        """Salva obras no banco de dados"""
        try:
            init_db()  # Garantir que as tabelas existam
            
            db = SessionLocal()
            try:
                # Criar uma sessão de scraping para este processo
                session = ScrapingSession(
                    scraper_name="obras_especificas",
                    status="concluido",
                    total_obras=len(obras),
                    inicio=datetime.utcnow(),
                    fim=datetime.utcnow()
                )
                db.add(session)
                db.commit()
                db.refresh(session)
                
                obras_salvas = 0
                obras_duplicadas = 0
                
                for obra_data in obras:
                    try:
                        url_obra = obra_data.get('url', '')
                        scraper_name = obra_data.get('scraper', '')
                        
                        if not url_obra or not scraper_name:
                            continue
                        
                        # Verificar se já existe
                        obra_existente = db.query(Obra).filter(
                            Obra.url == url_obra,
                            Obra.scraper_name == scraper_name
                        ).first()
                        
                        if obra_existente:
                            obras_duplicadas += 1
                            continue
                        
                        # Criar nova obra
                        obra = Obra(
                            session_id=session.id,
                            scraper_name=scraper_name,
                            categoria=obra_data.get('categoria', 'Quadros'),
                            nome_artista=obra_data.get('nome_artista'),
                            titulo=obra_data.get('titulo'),
                            descricao=obra_data.get('descricao'),
                            valor=obra_data.get('valor'),
                            valor_atualizado=obra_data.get('valor_atual'),
                            numero_lances=obra_data.get('numero_lances', 0),
                            lote=obra_data.get('lote'),
                            data_inicio_leilao=obra_data.get('data_inicio_leilao'),
                            data_leilao=obra_data.get('data_final_leilao'),
                            leiloeiro=obra_data.get('leiloeiro'),
                            local=obra_data.get('local'),
                            url=url_obra,
                            url_original=obra_data.get('url_original'),
                            data_coleta=datetime.utcnow()
                        )
                        db.add(obra)
                        obras_salvas += 1
                        
                        # Commit a cada 10 obras
                        if obras_salvas % 10 == 0:
                            db.commit()
                            print(f"  💾 {obras_salvas} obras salvas no banco...")
                        
                        # Mapear URL para ID após commit
                        db.refresh(obra)
                        url_para_monitorar = obra_data.get('url_original') or url_obra
                        self.obras_ids_banco[url_para_monitorar] = obra.id
                    except Exception as e:
                        print(f"  ⚠️ Erro ao salvar obra {obra_data.get('url', 'N/A')}: {e}")
                        continue
                
                # Commit final
                db.commit()
                
                # Mapear URLs das obras duplicadas também (para atualização futura)
                for obra_data in obras:
                    url_obra = obra_data.get('url', '')
                    scraper_name = obra_data.get('scraper', '')
                    if url_obra and scraper_name:
                        obra_existente = db.query(Obra).filter(
                            Obra.url == url_obra,
                            Obra.scraper_name == scraper_name
                        ).first()
                        if obra_existente:
                            url_para_monitorar = obra_data.get('url_original') or url_obra
                            self.obras_ids_banco[url_para_monitorar] = obra_existente.id
                
                print(f"\n✅ {obras_salvas} obras salvas no banco de dados")
                if obras_duplicadas > 0:
                    print(f"⏭️  {obras_duplicadas} obras duplicadas ignoradas")
                print(f"📋 {len(self.obras_ids_banco)} obras mapeadas para monitoramento")
                
            finally:
                db.close()
        except Exception as e:
            print(f"  ⚠️ Erro ao salvar obras no banco: {e}")
            import traceback
            traceback.print_exc()
    
    def exportar_para_excel(self, obras: List[Dict], filename: str = None):
        """Exporta obras e histórico de valores para Excel separados por site"""
        # Separar obras por scraper
        obras_iarremate = [o for o in obras if o.get('scraper') == 'iarremate']
        obras_leiloes_br = [o for o in obras if o.get('scraper') == 'leiloes_br']
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        arquivos_gerados = []
        
        # Criar diretório se não existir
        Path("output").mkdir(parents=True, exist_ok=True)
        
        # Função auxiliar para criar Excel
        def criar_excel(obras_lista: List[Dict], nome_arquivo: str, nome_site: str) -> str:
            wb = Workbook()
            wb.remove(wb.active)  # Remover planilha padrão
            
            # Planilha 1: Resumo das Obras
            ws_obras = wb.create_sheet("Obras")
            
            # Cabeçalhos
            headers = [
                'URL', 'Scraper', 'Título', 'Artista', 'Lote', 
                'Valor Atual', 'Nº Lances', 'Valor Inicial', 'Visitas', 'Status Lote',
                'Data Início Leilão', 'Data Final Leilão', 'Status Leilão', 'Leiloeiro', 'Local', 'Categoria'
            ]
            
            ws_obras.append(headers)
            
            # Estilizar cabeçalhos
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in ws_obras[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Adicionar dados
            for obra in obras_lista:
                row = [
                    obra.get('url', 'N/A'),
                    obra.get('scraper', 'N/A'),
                    obra.get('titulo', 'N/A'),
                    obra.get('nome_artista', 'N/A'),
                    obra.get('lote', 'N/A'),
                    obra.get('valor_atual', obra.get('valor', 'N/A')),
                    obra.get('numero_lances', 0),
                    obra.get('valor', 'N/A'),
                    obra.get('visitas', 'N/A'),
                    obra.get('status_lote', 'Disponível'),
                    obra.get('data_inicio_leilao', 'N/A'),
                    obra.get('data_final_leilao', 'N/A'),
                    obra.get('status_leilao', 'N/A'),
                    obra.get('leiloeiro', 'N/A'),
                    obra.get('local', 'N/A'),
                    obra.get('categoria', 'N/A')
                ]
                ws_obras.append(row)
            
            # Aplicar bordas e ajustar largura das colunas
            for row in ws_obras.iter_rows(min_row=2, max_row=ws_obras.max_row):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Ajustar largura das colunas
            column_widths = {
                'A': 50,  # URL
                'B': 12,  # Scraper
                'C': 40,  # Título
                'D': 25,  # Artista
                'E': 10,  # Lote
                'F': 15,  # Valor Atual
                'G': 12,  # Nº Lances
                'H': 15,  # Valor Inicial
                'I': 12,  # Visitas
                'J': 15,  # Status Lote
                'K': 20,  # Data Início
                'L': 20,  # Data Final
                'M': 15,  # Status Leilão
                'N': 20,  # Leiloeiro
                'O': 20,  # Local
                'P': 15   # Categoria
            }
            
            for col, width in column_widths.items():
                ws_obras.column_dimensions[col].width = width
            
            # Planilha 2: Histórico de Valores (apenas para obras deste site)
            ws_historico = wb.create_sheet("Histórico de Valores")
            
            # Cabeçalhos do histórico
            headers_historico = ['URL', 'Lance', 'Valor', 'Data/Hora', 'Mudança']
            ws_historico.append(headers_historico)
            
            # Estilizar cabeçalhos
            for cell in ws_historico[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Adicionar histórico de valores apenas para obras deste site
            for obra in obras_lista:
                url = obra.get('url_original') or obra.get('url', '')
                if url in self.historicos:
                    historico = self.historicos[url]
                    valores = historico.obter_todos_valores()
                    valor_anterior = None
                    
                    for valor, timestamp, numero_lance in valores:
                        mudanca = ''
                        if valor_anterior:
                            try:
                                valor_ant_num = float(valor_anterior.replace('.', '').replace(',', '.').replace('R$', '').strip())
                                valor_atual_num = float(valor.replace('.', '').replace(',', '.').replace('R$', '').strip())
                                diferenca = valor_atual_num - valor_ant_num
                                if diferenca > 0:
                                    mudanca = f"+R$ {diferenca:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                                elif diferenca < 0:
                                    mudanca = f"R$ {diferenca:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                            except:
                                pass
                        
                        row = [
                            url,
                            f"Lance {numero_lance}" if numero_lance > 0 else "Valor Inicial",
                            valor,
                            timestamp.strftime('%d/%m/%Y %H:%M:%S'),
                            mudanca
                        ]
                        ws_historico.append(row)
                        valor_anterior = valor
            
            # Aplicar bordas e ajustar largura
            for row in ws_historico.iter_rows(min_row=2, max_row=ws_historico.max_row):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical='top')
            
            column_widths_hist = {
                'A': 50,  # URL
                'B': 15,  # Lance
                'C': 15,  # Valor
                'D': 20,  # Data/Hora
                'E': 15   # Mudança
            }
            
            for col, width in column_widths_hist.items():
                ws_historico.column_dimensions[col].width = width
            
            # Salvar arquivo (com tratamento de erro se arquivo estiver aberto)
            tentativas = 3
            filename_original = nome_arquivo
            
            for tentativa in range(tentativas):
                try:
                    wb.save(nome_arquivo)
                    if tentativa > 0:
                        print(f"  ✅ Arquivo Excel salvo: {nome_arquivo}")
                    else:
                        print(f"  ✅ Arquivo Excel {nome_site} salvo: {nome_arquivo}")
                    return nome_arquivo
                except PermissionError:
                    if tentativa < tentativas - 1:
                        print(f"  ⚠️ Arquivo está aberto (tentativa {tentativa + 1}/{tentativas}). Aguardando...")
                        time.sleep(2)
                    else:
                        # Última tentativa: criar arquivo com nome único
                        timestamp_novo = datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:-3]
                        nome_arquivo = f"output/{nome_site.lower().replace(' ', '_')}_{timestamp_novo}.xlsx"
                        try:
                            wb.save(nome_arquivo)
                            print(f"  ✅ Arquivo Excel salvo (novo nome): {nome_arquivo}")
                            print(f"  ⚠️ O arquivo original ({filename_original}) estava aberto.")
                            return nome_arquivo
                        except Exception as e:
                            print(f"  ❌ Erro ao salvar Excel após {tentativas} tentativas: {e}")
                            return filename_original
                except Exception as e:
                    print(f"  ❌ Erro ao salvar Excel {nome_site}: {e}")
                    return filename_original
            
            return filename_original
        
        # Gerar Excel para iArremate
        if obras_iarremate:
            filename_iarremate = f"output/iarremate_{timestamp}.xlsx"
            arquivo_iarremate = criar_excel(obras_iarremate, filename_iarremate, "iArremate")
            arquivos_gerados.append(arquivo_iarremate)
        
        # Gerar Excel para LeilõesBR
        if obras_leiloes_br:
            filename_leiloes_br = f"output/leiloes_br_{timestamp}.xlsx"
            arquivo_leiloes_br = criar_excel(obras_leiloes_br, filename_leiloes_br, "LeilõesBR")
            arquivos_gerados.append(arquivo_leiloes_br)
        
        return arquivos_gerados


def main():
    """Função principal"""
    extrator = ExtratorObrasEspecificas()
    
    # Extrair todas as obras
    obras = extrator.extrair_todas_obras()
    
    if not obras:
        print("❌ Nenhuma obra foi extraída com sucesso.")
        return
    
    # Salvar obras no banco de dados
    print("\n" + "=" * 80)
    print("SALVANDO OBRAS NO BANCO DE DADOS")
    print("=" * 80)
    extrator.salvar_obras_no_banco(obras)
    
    # Iniciar monitoramento para obras com leilão ativo
    print("\n" + "=" * 80)
    print("INICIANDO MONITORAMENTO DE VALORES")
    print("=" * 80)
    
    for obra in obras:
        url = obra.get('url_original') or obra.get('url')
        extrator.iniciar_monitoramento(url, obra)
    
    # Exportar para Excel imediatamente (separados por site)
    print("\n" + "=" * 80)
    print("EXPORTANDO PARA EXCEL (SEPARADOS POR SITE)")
    print("=" * 80)
    
    if extrator.obras_puladas > 0:
        print(f"ℹ️  {extrator.obras_puladas} obras foram puladas por já existirem no banco de dados")
    
    arquivos_gerados = extrator.exportar_para_excel(obras)
    
    print("\n" + "=" * 80)
    print("MONITORAMENTO EM ANDAMENTO")
    print("=" * 80)
    print("O monitoramento continuará rodando em background.")
    print("Pressione Ctrl+C para parar e exportar novamente.")
    print("=" * 80)
    
    try:
        # Manter o script rodando para monitoramento contínuo
        ultima_exportacao = time.time()
        intervalo_exportacao = 300  # 5 minutos
        
        print("\n🔄 Monitoramento contínuo ativo!")
        print("   - Verificando valores a cada 30 segundos")
        print("   - Exportando Excel a cada 5 minutos")
        print("   - Pressione Ctrl+C para parar\n")
        
        while True:
            time.sleep(30)  # Verificar a cada 30 segundos
            
            # Atualizar obras com valores mais recentes
            obras_atualizadas = []
            for obra in obras:
                url = obra.get('url_original') or obra.get('url')
                if url in extrator.historicos:
                    ultimo_valor = extrator.historicos[url].obter_ultimo_valor()
                    if ultimo_valor:
                        obra['valor_atual'] = ultimo_valor[0]
                        obra['numero_lances'] = ultimo_valor[2] if len(ultimo_valor) > 2 else 0
                obras_atualizadas.append(obra)
            obras = obras_atualizadas
            
            # Exportar novamente a cada 5 minutos
            tempo_atual = time.time()
            if tempo_atual - ultima_exportacao >= intervalo_exportacao:
                print(f"\n[{datetime.now().strftime('%H:%M:%S')}] 📊 Re-exportando Excel com valores atualizados...")
                try:
                    novos_arquivos = extrator.exportar_para_excel(obras)
                    # Atualizar lista de arquivos gerados
                    if novos_arquivos:
                        arquivos_gerados = novos_arquivos
                        print(f"  ℹ️ Arquivos atualizados: {', '.join(arquivos_gerados)}")
                except Exception as e:
                    print(f"  ⚠️ Erro ao exportar Excel: {e}")
                    print(f"  💡 Feche o arquivo Excel se estiver aberto e aguarde próxima exportação.")
                ultima_exportacao = tempo_atual
    except KeyboardInterrupt:
        print("\n\n" + "=" * 80)
        print("PARANDO MONITORAMENTO")
        print("=" * 80)
        
        # Parar todos os monitores
        for url in list(extrator.monitores_ativos.keys()):
            extrator.parar_monitoramento(url)
        
        # Aguardar threads finalizarem
        time.sleep(2)
        
        # Exportar final
        print("\nExportando versão final do Excel...")
        extrator.exportar_para_excel(obras)
        
        print("\n✅ Processo finalizado!")


if __name__ == "__main__":
    main()

